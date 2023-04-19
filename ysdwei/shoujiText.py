from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver import ActionChains
import re
import openpyxl
import  os
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook


import time
#打开空白浏览器
driver = webdriver.Chrome()
#访问网站
driver.get("http://www.chinaclear.cn/zdjs/flfg/law_index_new.shtml")

get_title = driver.find_elements(By.XPATH,'//div[text()="业务规则"]/following-sibling::div/div[@class="pageTabContent"]/ul/li/a')
get_time = driver.find_elements(By.XPATH,'//div[text()="业务规则"]/following-sibling::div/div[@class="pageTabContent"]/ul/li/span')

get_title_data = []
get_time_data = []

for i in get_title:
    # print(i.text)
    get_title_data.append(i.text)
del get_title_data[4:]
print(get_title_data)
for i in get_time:
    # print(i.text)
    get_time_data.append(i.text)
del get_time_data[4:]
print(get_time_data)
q = 8
t = 8
for i in range(2,6):

    get_title_dianji = driver.find_element(By.XPATH,'//div[text()="业务规则"]/following::ul/li[{}]'.format(i))
    get_title_dianji.click()
    get_title_two = driver.find_elements(By.XPATH,'//div[text()="业务规则"]/following-sibling::div/div[@class="pageTabContent"]/ul/li/a')
    get_time_two = driver.find_elements(By.XPATH,'//div[text()="业务规则"]/following-sibling::div/div[@class="pageTabContent"]/ul/li/span')
    for i in get_title_two:
        # print(i.text)

        # get_title_data.remove('')
        # for i in get_title_data:
        if i.text != '':
            get_title_data.append(i.text)

    del get_title_data[q:]
    q +=4
    print(get_title_data)
    for i in get_time_two:
        # print(i.text)

    # for i in get_time_data:
        if i.text != '':
            get_time_data.append(i.text)

    del get_time_data[t:]
    t += 4
    print(get_time_data)

excelPath = os.path.join(os.getcwd())
nameTime = time.strftime('%Y-%m-%d_%H-%M-%S')
excelName = 'Excel' + nameTime + '.xlsx'
ExcelFullName= os.path.join(excelPath,excelName)
wb = Workbook()
ws = wb.active
tableTitle = ['标题', '时间']
for col in range(len(tableTitle)):
    c = col + 1
    ws.cell(row=1, column=c).value = tableTitle[col]
for i in range(2,len(get_title_data)+ 1):
    ws.cell(i,1,get_title_data[i - 1])
wb.save(filename=ExcelFullName)
for i in range(2,len(get_time_data)+ 1):
    ws.cell(i,2,get_time_data[i - 1])
wb.save(filename=ExcelFullName)
driver.close()